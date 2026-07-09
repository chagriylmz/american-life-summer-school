from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re
import unicodedata

import openpyxl


PROJECT_ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = PROJECT_ROOT / "yazokulu.xlsx"
OUTPUT_PATH = PROJECT_ROOT / "sql" / "import_yazokulu_generated.sql"
SUMMARY_PATH = PROJECT_ROOT / "sql" / "import_yazokulu_summary.txt"

TARGET_SHEET_MARKER = "YAZ OKULU 1 TEMMUZ"

SOURCE_TO_IMPORT_TIMES = {
    "14:00-16:20": ("13:00", "15:20"),
    "16:30-18:50": ("15:30", "17:50"),
    "13:00-15:20": ("13:00", "15:20"),
    "15:30-17:50": ("15:30", "17:50"),
}

CLASS_GROUPS = [
    {"index_col": 1, "name_col": 2, "birth_col": 3, "notes_col": 4},
    {"index_col": 6, "name_col": 7, "birth_col": 8, "notes_col": 9},
    {"index_col": 11, "name_col": 12, "birth_col": 13, "notes_col": 14},
]


@dataclass(frozen=True)
class Block:
    header_row: int
    source_label: str
    source_start: str
    source_end: str
    import_start: str
    import_end: str
    teacher_row: int
    header_row_for_students: int
    data_start_row: int
    data_end_row: int


@dataclass(frozen=True)
class ClassRecord:
    class_key: str
    teacher_employee_code: str
    teacher_name: str
    room: str
    source_time_label: str
    import_start: str
    import_end: str


@dataclass(frozen=True)
class StudentRecord:
    student_code: str
    full_name: str
    birth_year: int | None
    phone: str | None
    notes: str | None


@dataclass(frozen=True)
class EnrollmentRecord:
    class_key: str
    student_code: str


def normalize_text(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def slugify(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    ascii_text = normalized.encode("ascii", "ignore").decode("ascii")
    slug = re.sub(r"[^A-Za-z0-9]+", "-", ascii_text).strip("-").upper()
    return slug or "UNKNOWN"


def sql_string(value: object | None) -> str:
    if value is None:
        return "null"
    return "'" + str(value).replace("'", "''") + "'"


def sql_int(value: int | None) -> str:
    return "null" if value is None else str(value)


def parse_time_label(value: str) -> tuple[str, str, str] | None:
    match = re.search(r"(\d{1,2})[:.](\d{2})\s*[-–]\s*(\d{1,2})[:.](\d{2})", value)
    if not match:
        return None
    start = f"{int(match.group(1)):02d}:{match.group(2)}"
    end = f"{int(match.group(3)):02d}:{match.group(4)}"
    return f"{start}-{end}", start, end


def parse_teacher_room(value: str) -> tuple[str, str] | None:
    cleaned = normalize_text(value)
    match = re.match(r"^(.+?)[\s-]+(\d{2,4})$", cleaned)
    if not match:
        return None
    return normalize_text(match.group(1)).upper(), match.group(2)


def parse_birth_year(value: object) -> int | None:
    if value is None:
        return None
    if isinstance(value, int):
        return value if 1900 <= value <= 2100 else None
    if isinstance(value, float):
        year = int(value)
        return year if 1900 <= year <= 2100 else None
    match = re.search(r"(19|20)\d{2}", str(value))
    return int(match.group(0)) if match else None


def looks_like_student_name(value: str) -> bool:
    if len(value) < 3:
        return False
    lowered = value.casefold()
    blocked = {"ok", "iade", "yarın", "belli değil", "gelmeyecek", "gelmiyor"}
    return lowered not in blocked and bool(re.search(r"[A-Za-zÇĞİÖŞÜçğıöşü]", value))


def detect_phone_column(ws, header_row: int, group: dict[str, int]) -> int | None:
    for col in range(group["name_col"], min(group["name_col"] + 5, ws.max_column) + 1):
        header = normalize_text(ws.cell(header_row, col).value).casefold()
        if any(marker in header for marker in ("telefon", "tel", "gsm", "phone")):
            return col
    return None


def find_target_sheet(wb) -> str:
    matches = [name for name in wb.sheetnames if TARGET_SHEET_MARKER in name]
    if len(matches) == 1:
        return matches[0]
    yaz_okulu_matches = [name for name in wb.sheetnames if "YAZ OKULU" in name.upper()]
    if len(yaz_okulu_matches) == 1:
        return yaz_okulu_matches[0]
    raise RuntimeError(
        "Could not uniquely identify the summer school worksheet. "
        f"Available sheets: {', '.join(wb.sheetnames)}"
    )


def detect_blocks(ws) -> list[Block]:
    time_rows: list[tuple[int, str, str, str, str, str]] = []
    for row in range(1, ws.max_row + 1):
        row_text = " ".join(
            normalize_text(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)
        )
        parsed = parse_time_label(row_text)
        if not parsed:
            continue
        source_label, source_start, source_end = parsed
        if source_label not in SOURCE_TO_IMPORT_TIMES:
            continue
        import_start, import_end = SOURCE_TO_IMPORT_TIMES[source_label]
        time_rows.append((row, source_label, source_start, source_end, import_start, import_end))

    blocks: list[Block] = []
    for index, item in enumerate(time_rows):
        next_time_row = time_rows[index + 1][0] if index + 1 < len(time_rows) else ws.max_row + 1
        header_row, source_label, source_start, source_end, import_start, import_end = item
        blocks.append(
            Block(
                header_row=header_row,
                source_label=source_label,
                source_start=source_start,
                source_end=source_end,
                import_start=import_start,
                import_end=import_end,
                teacher_row=header_row + 2,
                header_row_for_students=header_row + 3,
                data_start_row=header_row + 4,
                data_end_row=next_time_row - 1,
            )
        )
    return blocks


def collect_records(ws, blocks: list[Block]):
    classes: dict[str, ClassRecord] = {}
    students: dict[str, StudentRecord] = {}
    enrollments: set[EnrollmentRecord] = set()

    for block in blocks:
        for group in CLASS_GROUPS:
            teacher_room = parse_teacher_room(ws.cell(block.teacher_row, group["name_col"]).value)
            if not teacher_room:
                continue
            teacher_name, room = teacher_room
            class_key = f"SUMMER-{block.import_start}-{teacher_name}-ROOM-{room}"
            classes[class_key] = ClassRecord(
                class_key=class_key,
                teacher_employee_code=f"YAZ-{slugify(teacher_name)}",
                teacher_name=teacher_name,
                room=room,
                source_time_label=block.source_label,
                import_start=block.import_start,
                import_end=block.import_end,
            )

            phone_col = detect_phone_column(ws, block.header_row_for_students, group)
            for row in range(block.data_start_row, block.data_end_row + 1):
                full_name = normalize_text(ws.cell(row, group["name_col"]).value)
                if not looks_like_student_name(full_name):
                    continue

                birth_year = parse_birth_year(ws.cell(row, group["birth_col"]).value)
                phone = normalize_text(ws.cell(row, phone_col).value) if phone_col else ""
                notes = normalize_text(ws.cell(row, group["notes_col"]).value)
                student_code = f"YAZ-{slugify(full_name)}-{birth_year or 'NOYEAR'}"

                students[student_code] = StudentRecord(
                    student_code=student_code,
                    full_name=full_name,
                    birth_year=birth_year,
                    phone=phone or None,
                    notes=notes or None,
                )
                enrollments.add(EnrollmentRecord(class_key=class_key, student_code=student_code))

    return sorted(classes.values(), key=lambda item: item.class_key), sorted(
        students.values(), key=lambda item: item.student_code
    ), sorted(enrollments, key=lambda item: (item.class_key, item.student_code))


def values_rows(rows: list[tuple[str, ...]]) -> str:
    return ",\n".join("  (" + ", ".join(row) + ")" for row in rows)


def build_sql(classes: list[ClassRecord], students: list[StudentRecord], enrollments: list[EnrollmentRecord]) -> str:
    teacher_names = sorted({item.teacher_name for item in classes})
    teacher_rows = [
        (
            sql_string(f"YAZ-{slugify(name)}"),
            sql_string(name),
            sql_string(f"Imported from {WORKBOOK_PATH.name}"),
        )
        for name in teacher_names
    ]
    class_rows = [
        (
            sql_string(item.class_key),
            sql_string(item.teacher_employee_code),
            sql_string(item.teacher_name),
            sql_string(item.room),
            sql_string(item.source_time_label),
            sql_string(item.import_start),
            sql_string(item.import_end),
        )
        for item in classes
    ]
    student_rows = [
        (
            sql_string(item.student_code),
            sql_string(item.full_name),
            sql_int(item.birth_year),
            sql_string(item.phone),
            sql_string(item.notes),
        )
        for item in students
    ]
    enrollment_rows = [
        (sql_string(item.class_key), sql_string(item.student_code)) for item in enrollments
    ]

    return f"""-- Generated from yazokulu.xlsx for American Life Summer School.
-- Source worksheet: YAZ OKULU 1 TEMMUZ-1 AĞUSTOS 20
-- Safe to re-run. Teachers/students are upserted by stable codes; classes are matched by deterministic names.
--
-- Important:
-- The source worksheet displays 14:00-16:20 and 16:30-18:50.
-- This import maps them to the requested app sessions 13:00-15:20 and 15:30-17:50.

begin;

create temp table tmp_yazokulu_teachers (
  employee_code text primary key,
  display_name text not null,
  bio text
) on commit drop;

create temp table tmp_yazokulu_classes (
  class_key text primary key,
  teacher_employee_code text not null,
  teacher_name text not null,
  room text not null,
  source_time_label text not null,
  starts_at time not null,
  ends_at time not null
) on commit drop;

create temp table tmp_yazokulu_students (
  student_code text primary key,
  full_name text not null,
  birth_year integer,
  phone text,
  notes text
) on commit drop;

create temp table tmp_yazokulu_enrollments (
  class_key text not null,
  student_code text not null,
  primary key (class_key, student_code)
) on commit drop;

insert into tmp_yazokulu_teachers (employee_code, display_name, bio)
values
{values_rows(teacher_rows)};

insert into tmp_yazokulu_classes (class_key, teacher_employee_code, teacher_name, room, source_time_label, starts_at, ends_at)
values
{values_rows(class_rows)};

insert into tmp_yazokulu_students (student_code, full_name, birth_year, phone, notes)
values
{values_rows(student_rows)};

insert into tmp_yazokulu_enrollments (class_key, student_code)
values
{values_rows(enrollment_rows)};

insert into public.teachers (
  employee_code,
  display_name,
  bio,
  languages,
  is_active,
  hired_at
)
select
  employee_code,
  display_name,
  bio,
  array['English'],
  true,
  current_date
from tmp_yazokulu_teachers
on conflict (employee_code) do update
set
  display_name = excluded.display_name,
  bio = excluded.bio,
  languages = excluded.languages,
  is_active = excluded.is_active;

insert into public.students (
  student_code,
  full_name,
  phone,
  date_of_birth,
  native_language,
  target_language,
  current_level,
  enrollment_date,
  is_active,
  notes
)
select
  student_code,
  full_name,
  phone,
  case when birth_year is null then null else make_date(birth_year, 1, 1) end,
  'Turkish',
  'English',
  'beginner',
  current_date,
  true,
  concat_ws('; ', 'Imported from yazokulu.xlsx', case when birth_year is null then null else 'Birth year: ' || birth_year::text end, notes)
from tmp_yazokulu_students
on conflict (student_code) do update
set
  full_name = excluded.full_name,
  phone = excluded.phone,
  date_of_birth = excluded.date_of_birth,
  native_language = excluded.native_language,
  target_language = excluded.target_language,
  current_level = excluded.current_level,
  is_active = excluded.is_active,
  notes = excluded.notes;

create temp table tmp_yazokulu_class_ids (
  class_key text primary key,
  class_id uuid not null
) on commit drop;

do $$
declare
  row_data record;
  teacher_uuid uuid;
  existing_class_id uuid;
  class_name text;
begin
  for row_data in
    select
      c.*,
      t.id as teacher_id
    from tmp_yazokulu_classes c
    join public.teachers t on t.employee_code = c.teacher_employee_code
  loop
    teacher_uuid := row_data.teacher_id;
    class_name := 'American Life Summer School - ' || to_char(row_data.starts_at, 'HH24:MI') || '-' || to_char(row_data.ends_at, 'HH24:MI') || ' - ' || row_data.teacher_name || ' - Room ' || row_data.room;

    select id
    into existing_class_id
    from public.classes
    where name = class_name
      and teacher_id = teacher_uuid
    order by created_at desc
    limit 1;

    if existing_class_id is null then
      insert into public.classes (
        teacher_id,
        name,
        language,
        level,
        status,
        capacity,
        start_date,
        end_date,
        schedule,
        location,
        meeting_url
      )
      values (
        teacher_uuid,
        class_name,
        'English',
        'beginner',
        'active',
        16,
        current_date,
        current_date + interval '8 weeks',
        jsonb_build_array(jsonb_build_object(
          'days', 'Monday-Thursday',
          'starts_at', to_char(row_data.starts_at, 'HH24:MI'),
          'ends_at', to_char(row_data.ends_at, 'HH24:MI'),
          'source_time', row_data.source_time_label
        )),
        'Room ' || row_data.room,
        null
      )
      returning id into existing_class_id;
    else
      update public.classes
      set
        language = 'English',
        level = 'beginner',
        status = 'active',
        capacity = 16,
        start_date = current_date,
        end_date = current_date + interval '8 weeks',
        schedule = jsonb_build_array(jsonb_build_object(
          'days', 'Monday-Thursday',
          'starts_at', to_char(row_data.starts_at, 'HH24:MI'),
          'ends_at', to_char(row_data.ends_at, 'HH24:MI'),
          'source_time', row_data.source_time_label
        )),
        location = 'Room ' || row_data.room,
        meeting_url = null
      where id = existing_class_id;
    end if;

    insert into tmp_yazokulu_class_ids (class_key, class_id)
    values (row_data.class_key, existing_class_id)
    on conflict (class_key) do update set class_id = excluded.class_id;

    insert into public.lessons (
      class_id,
      teacher_id,
      lesson_date,
      starts_at,
      ends_at,
      title,
      objectives,
      materials,
      homework,
      status
    )
    values (
      existing_class_id,
      teacher_uuid,
      current_date,
      row_data.starts_at,
      row_data.ends_at,
      'Summer School Session - ' || row_data.teacher_name || ' Room ' || row_data.room,
      'Imported session from yazokulu.xlsx.',
      null,
      null,
      'scheduled'
    )
    on conflict (class_id, lesson_date, starts_at) do update
    set
      teacher_id = excluded.teacher_id,
      ends_at = excluded.ends_at,
      title = excluded.title,
      objectives = excluded.objectives,
      status = excluded.status;
  end loop;
end $$;

insert into public.class_students (
  class_id,
  student_id,
  status,
  joined_at
)
select
  c.class_id,
  s.id,
  'active',
  current_date
from tmp_yazokulu_enrollments e
join tmp_yazokulu_class_ids c on c.class_key = e.class_key
join public.students s on s.student_code = e.student_code
on conflict on constraint class_students_pkey do update
set
  status = excluded.status,
  joined_at = excluded.joined_at,
  left_at = null;

commit;
"""


def build_self_contained_sql(
    classes: list[ClassRecord], students: list[StudentRecord], enrollments: list[EnrollmentRecord]
) -> str:
    teacher_names = sorted({item.teacher_name for item in classes})
    teacher_rows = [
        (
            sql_string(f"YAZ-{slugify(name)}"),
            sql_string(name),
            sql_string(f"Imported from {WORKBOOK_PATH.name}"),
        )
        for name in teacher_names
    ]
    class_rows = [
        (
            sql_string(item.class_key),
            sql_string(item.teacher_employee_code),
            sql_string(item.teacher_name),
            sql_string(item.room),
            sql_string(item.source_time_label),
            sql_string(item.import_start),
            sql_string(item.import_end),
        )
        for item in classes
    ]
    student_rows = [
        (
            sql_string(item.student_code),
            sql_string(item.full_name),
            sql_int(item.birth_year),
            sql_string(item.phone),
            sql_string(item.notes),
        )
        for item in students
    ]
    enrollment_rows = [
        (sql_string(item.class_key), sql_string(item.student_code)) for item in enrollments
    ]

    return f"""-- Generated from yazokulu.xlsx for American Life Summer School.
-- Source worksheet: YAZ OKULU 1 TEMMUZ-1 AĞUSTOS 20
-- Safe to re-run. Teachers/students are upserted by stable codes; classes are matched by deterministic names.
-- Fully self-contained: this file does not rely on temporary or staging tables.
--
-- Important:
-- The source worksheet displays 14:00-16:20 and 16:30-18:50.
-- This import maps them to the requested app sessions 13:00-15:20 and 15:30-17:50.

begin;

with teacher_data(employee_code, display_name, bio) as (
  values
{values_rows(teacher_rows)}
)
insert into public.teachers (
  employee_code,
  display_name,
  bio,
  languages,
  is_active,
  hired_at
)
select
  employee_code,
  display_name,
  bio,
  array['English'],
  true,
  current_date
from teacher_data
on conflict (employee_code) do update
set
  display_name = excluded.display_name,
  bio = excluded.bio,
  languages = excluded.languages,
  is_active = excluded.is_active;

with student_data(student_code, full_name, birth_year, phone, notes) as (
  values
{values_rows(student_rows)}
)
insert into public.students (
  student_code,
  full_name,
  phone,
  date_of_birth,
  native_language,
  target_language,
  current_level,
  enrollment_date,
  is_active,
  notes
)
select
  student_code,
  full_name,
  phone,
  case when birth_year is null then null else make_date(birth_year, 1, 1) end,
  'Turkish',
  'English',
  'beginner',
  current_date,
  true,
  concat_ws('; ', 'Imported from yazokulu.xlsx', case when birth_year is null then null else 'Birth year: ' || birth_year::text end, notes)
from student_data
on conflict (student_code) do update
set
  full_name = excluded.full_name,
  phone = excluded.phone,
  date_of_birth = excluded.date_of_birth,
  native_language = excluded.native_language,
  target_language = excluded.target_language,
  current_level = excluded.current_level,
  is_active = excluded.is_active,
  notes = excluded.notes;

do $import$
declare
  row_data record;
  teacher_uuid uuid;
  existing_class_id uuid;
  class_name text;
begin
  for row_data in
    select
      c.class_key,
      c.teacher_employee_code,
      c.teacher_name,
      c.room,
      c.source_time_label,
      c.starts_at::time as starts_at,
      c.ends_at::time as ends_at,
      t.id as teacher_id
    from (
      values
{values_rows(class_rows)}
    ) as c(class_key, teacher_employee_code, teacher_name, room, source_time_label, starts_at, ends_at)
    join public.teachers t on t.employee_code = c.teacher_employee_code
  loop
    teacher_uuid := row_data.teacher_id;
    class_name := 'American Life Summer School - ' || to_char(row_data.starts_at, 'HH24:MI') || '-' || to_char(row_data.ends_at, 'HH24:MI') || ' - ' || row_data.teacher_name || ' - Room ' || row_data.room;

    select id
    into existing_class_id
    from public.classes
    where name = class_name
      and teacher_id = teacher_uuid
    order by created_at desc
    limit 1;

    if existing_class_id is null then
      insert into public.classes (
        teacher_id,
        name,
        language,
        level,
        status,
        capacity,
        start_date,
        end_date,
        schedule,
        location,
        meeting_url
      )
      values (
        teacher_uuid,
        class_name,
        'English',
        'beginner',
        'active',
        16,
        current_date,
        current_date + interval '8 weeks',
        jsonb_build_array(jsonb_build_object(
          'days', 'Monday-Thursday',
          'starts_at', to_char(row_data.starts_at, 'HH24:MI'),
          'ends_at', to_char(row_data.ends_at, 'HH24:MI'),
          'source_time', row_data.source_time_label
        )),
        'Room ' || row_data.room,
        null
      )
      returning id into existing_class_id;
    else
      update public.classes
      set
        language = 'English',
        level = 'beginner',
        status = 'active',
        capacity = 16,
        start_date = current_date,
        end_date = current_date + interval '8 weeks',
        schedule = jsonb_build_array(jsonb_build_object(
          'days', 'Monday-Thursday',
          'starts_at', to_char(row_data.starts_at, 'HH24:MI'),
          'ends_at', to_char(row_data.ends_at, 'HH24:MI'),
          'source_time', row_data.source_time_label
        )),
        location = 'Room ' || row_data.room,
        meeting_url = null
      where id = existing_class_id;
    end if;

    insert into public.lessons (
      class_id,
      teacher_id,
      lesson_date,
      starts_at,
      ends_at,
      title,
      objectives,
      materials,
      homework,
      status
    )
    values (
      existing_class_id,
      teacher_uuid,
      current_date,
      row_data.starts_at,
      row_data.ends_at,
      'Summer School Session - ' || row_data.teacher_name || ' Room ' || row_data.room,
      'Imported session from yazokulu.xlsx.',
      null,
      null,
      'scheduled'
    )
    on conflict (class_id, lesson_date, starts_at) do update
    set
      teacher_id = excluded.teacher_id,
      ends_at = excluded.ends_at,
      title = excluded.title,
      objectives = excluded.objectives,
      status = excluded.status;
  end loop;
end $import$;

with
  enrollment_data(class_key, student_code) as (
    values
{values_rows(enrollment_rows)}
  ),
  class_data(class_key, teacher_employee_code, teacher_name, room, source_time_label, starts_at, ends_at) as (
    values
{values_rows(class_rows)}
  ),
  resolved_classes as (
    select
      c.class_key,
      cls.id as class_id
    from class_data c
    join public.teachers t on t.employee_code = c.teacher_employee_code
    join public.classes cls
      on cls.teacher_id = t.id
      and cls.name = 'American Life Summer School - ' || to_char(c.starts_at::time, 'HH24:MI') || '-' || to_char(c.ends_at::time, 'HH24:MI') || ' - ' || c.teacher_name || ' - Room ' || c.room
  )
insert into public.class_students (
  class_id,
  student_id,
  status,
  joined_at
)
select
  rc.class_id,
  s.id,
  'active',
  current_date
from enrollment_data e
join resolved_classes rc on rc.class_key = e.class_key
join public.students s on s.student_code = e.student_code
on conflict on constraint class_students_pkey do update
set
  status = excluded.status,
  joined_at = excluded.joined_at,
  left_at = null;

commit;
"""


def main() -> None:
    wb = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
    sheet_name = find_target_sheet(wb)
    ws = wb[sheet_name]
    blocks = detect_blocks(ws)
    classes, students, enrollments = collect_records(ws, blocks)

    if len(blocks) != 2:
        raise RuntimeError(f"Expected 2 time blocks, detected {len(blocks)}.")
    if not classes:
        raise RuntimeError("No classes detected.")
    if not students:
        raise RuntimeError("No students detected.")

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(build_self_contained_sql(classes, students, enrollments), encoding="utf-8")

    summary = [
        f"Workbook: {WORKBOOK_PATH.name}",
        f"Worksheet: {sheet_name}",
        f"Detected time blocks: {len(blocks)}",
        f"Detected class/session groups: {len(classes)}",
        f"Detected unique students: {len(students)}",
        f"Detected enrollments: {len(enrollments)}",
        "",
        "Classes:",
    ]
    for item in classes:
        summary.append(
            f"- {item.import_start}-{item.import_end} | {item.teacher_name} | Room {item.room} | source {item.source_time_label}"
        )
    SUMMARY_PATH.write_text("\n".join(summary) + "\n", encoding="utf-8")

    print("\n".join(summary))
    print(f"\nGenerated: {OUTPUT_PATH}")
    print(f"Summary: {SUMMARY_PATH}")


if __name__ == "__main__":
    main()
