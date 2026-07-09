from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
import json
import re

import openpyxl


PROJECT_ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = PROJECT_ROOT / "yazokulu.xlsx"
OUTPUT_DIR = PROJECT_ROOT / "outputs"
AUDIT_TXT = OUTPUT_DIR / "yazokulu_workbook_audit.txt"
AUDIT_JSON = OUTPUT_DIR / "yazokulu_workbook_audit.json"
PREVIEW_TXT = OUTPUT_DIR / "yazokulu_import_preview.txt"
PREVIEW_JSON = OUTPUT_DIR / "yazokulu_import_preview.json"

TARGET_VISIBLE_SHEET_PREFIX = "YAZ OKULU-SINIF PLANLANMASI(TAS"
TIME_RE = re.compile(r"(\d{1,2})[:.](\d{2})\s*[-–]\s*(\d{1,2})[:.](\d{2})")


@dataclass(frozen=True)
class ClassGroup:
    teacher: str
    room: str
    session_time: str
    student_count: int


def clean(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def used_range(ws) -> str:
    non_empty_rows = []
    non_empty_cols = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and clean(cell.value):
                non_empty_rows.append(cell.row)
                non_empty_cols.append(cell.column)

    if not non_empty_rows:
        return "empty"

    start = ws.cell(min(non_empty_rows), min(non_empty_cols)).coordinate
    end = ws.cell(max(non_empty_rows), max(non_empty_cols)).coordinate
    return f"{start}:{end}"


def normalize_time(value: str) -> str | None:
    match = TIME_RE.search(value)
    if not match:
        return None
    return f"{int(match.group(1)):02d}:{match.group(2)}-{int(match.group(3)):02d}:{match.group(4)}"


def parse_teacher_room(value: object) -> tuple[str, str] | None:
    text = clean(value)
    teacher = ""
    room = ""

    teacher_first = re.match(r"^(.+?)[\s-]+(\d{2,4})$", text)
    room_first = re.match(r"^(\d{2,4})\s*-?\s*(.+?)$", text)

    if teacher_first:
        teacher = clean(teacher_first.group(1)).upper()
        room = teacher_first.group(2)
    elif room_first:
        room = room_first.group(1)
        teacher = clean(room_first.group(2)).upper()
    else:
        return None

    if not re.search(r"[A-Za-zÇĞİÖŞÜçğıöşü]", teacher):
        return None
    if teacher.casefold() in {"materyal", "öğrenci adı", "doğum tarihi"}:
        return None
    return teacher, room


def looks_like_student_name(value: object) -> bool:
    text = clean(value)
    if len(text) < 3:
        return False
    lowered = text.casefold()
    if lowered in {"ok", "iade", "yarın", "belli değil", "gelmeyecek", "gelmiyor"}:
        return False
    return bool(re.search(r"[A-Za-zÇĞİÖŞÜçğıöşü]", text))


def detected_session_times(ws) -> list[dict[str, object]]:
    matches = []
    for row_idx in range(1, ws.max_row + 1):
        row_text = " ".join(clean(ws.cell(row_idx, col).value) for col in range(1, ws.max_column + 1))
        time_value = normalize_time(row_text)
        if time_value:
            matches.append({"row": row_idx, "time": time_value, "text": row_text})
    return matches


def detected_teacher_names(ws) -> list[str]:
    teachers = set()
    for row in ws.iter_rows():
        for cell in row:
            parsed = parse_teacher_room(cell.value)
            if parsed:
                teachers.add(parsed[0])
    return sorted(teachers)


def find_target_sheet(wb):
    matches = [
        ws
        for ws in wb.worksheets
        if ws.sheet_state == "visible" and ws.title.startswith(TARGET_VISIBLE_SHEET_PREFIX)
    ]
    if len(matches) != 1:
        visible = [ws.title for ws in wb.worksheets if ws.sheet_state == "visible"]
        raise RuntimeError(
            f"Expected exactly one visible sheet starting with {TARGET_VISIBLE_SHEET_PREFIX!r}. "
            f"Found {len(matches)}. Visible sheets: {visible}"
        )
    return matches[0]


def find_preview_blocks(ws) -> list[dict[str, object]]:
    time_rows = detected_session_times(ws)
    blocks = []

    for index, item in enumerate(time_rows):
        header_row = int(item["row"])
        next_header_row = int(time_rows[index + 1]["row"]) if index + 1 < len(time_rows) else ws.max_row + 1
        teacher_row = header_row + 3
        student_header_row = header_row + 4
        data_start_row = header_row + 5
        data_end_row = next_header_row - 1

        for col in range(1, ws.max_column + 1):
            parsed = parse_teacher_room(ws.cell(teacher_row, col).value)
            if not parsed:
                continue
            teacher, room = parsed
            name_col = col + 1

            header_text = clean(ws.cell(student_header_row, name_col).value).casefold()
            if (
                "öğrenci" not in header_text
                and "ogrenci" not in header_text
                and "ad/soyad" not in header_text
            ):
                continue

            count = 0
            for row_idx in range(data_start_row, data_end_row + 1):
                if looks_like_student_name(ws.cell(row_idx, name_col).value):
                    count += 1

            blocks.append(
                {
                    "teacher": teacher,
                    "room": room,
                    "session_time": item["time"],
                    "student_count": count,
                    "source_header": item["text"],
                }
            )

    return blocks


def main() -> None:
    wb = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    audit = []
    for ws in wb.worksheets:
        audit.append(
            {
                "sheet_name": ws.title,
                "visibility": ws.sheet_state,
                "used_range": used_range(ws),
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "detected_teacher_names": detected_teacher_names(ws),
                "detected_session_times": detected_session_times(ws),
            }
        )

    target_ws = find_target_sheet(wb)
    preview_classes = find_preview_blocks(target_ws)
    unique_students_estimate = sum(item["student_count"] for item in preview_classes)

    preview = {
        "sheet_name": target_ws.title,
        "visibility": target_ws.sheet_state,
        "classes": preview_classes,
        "teachers": sorted({item["teacher"] for item in preview_classes}),
        "rooms": sorted({item["room"] for item in preview_classes}),
        "sessions": sorted({item["session_time"] for item in preview_classes}),
        "total_students": unique_students_estimate,
    }

    AUDIT_JSON.write_text(json.dumps(audit, ensure_ascii=False, indent=2), encoding="utf-8")
    PREVIEW_JSON.write_text(json.dumps(preview, ensure_ascii=False, indent=2), encoding="utf-8")

    audit_lines = ["Yaz Okulu Workbook Audit", ""]
    for sheet in audit:
        audit_lines.extend(
            [
                f"Sheet: {sheet['sheet_name']}",
                f"Visibility: {sheet['visibility']}",
                f"Range: {sheet['used_range']} ({sheet['max_row']} rows x {sheet['max_column']} columns)",
                f"Teachers: {', '.join(sheet['detected_teacher_names']) or '-'}",
                "Session times:",
            ]
        )
        if sheet["detected_session_times"]:
            for session in sheet["detected_session_times"]:
                audit_lines.append(f"  - row {session['row']}: {session['time']} | {session['text']}")
        else:
            audit_lines.append("  - -")
        audit_lines.append("")
    AUDIT_TXT.write_text("\n".join(audit_lines), encoding="utf-8")

    grouped = defaultdict(list)
    for item in preview_classes:
        grouped[item["session_time"]].append(item)

    preview_lines = [
        "Yaz Okulu Import Preview",
        "",
        f"Source sheet: {preview['sheet_name']}",
        f"Visibility: {preview['visibility']}",
        f"Teachers: {', '.join(preview['teachers']) or '-'}",
        f"Rooms: {', '.join(preview['rooms']) or '-'}",
        f"Sessions: {', '.join(preview['sessions']) or '-'}",
        f"Total students: {preview['total_students']}",
        "",
        "Classes:",
    ]

    for session_time in sorted(grouped):
        preview_lines.append(f"Session {session_time}")
        for item in sorted(grouped[session_time], key=lambda row: (row["teacher"], row["room"])):
            preview_lines.append(
                f"  - {item['teacher']} | Room {item['room']} | {item['student_count']} students"
            )
    PREVIEW_TXT.write_text("\n".join(preview_lines), encoding="utf-8")

    print("\n".join(preview_lines))
    print("")
    print(f"Audit written to: {AUDIT_TXT}")
    print(f"Preview written to: {PREVIEW_TXT}")


if __name__ == "__main__":
    main()
