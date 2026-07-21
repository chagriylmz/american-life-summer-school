from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re
import unicodedata

import openpyxl


PROJECT_ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = PROJECT_ROOT / "yazokulu.xlsx"
OUTPUT_PATH = PROJECT_ROOT / "sql" / "import_visible_yazokulu_approved.sql"
SUMMARY_PATH = PROJECT_ROOT / "sql" / "import_visible_yazokulu_approved_summary.txt"
TARGET_VISIBLE_SHEET_PREFIX = "YAZ OKULU-SINIF PLANLANMASI(TAS"


@dataclass(frozen=True)
class ClassRecord:
    class_key: str
    teacher_code: str
    teacher_name: str
    room: str
    starts_at: str
    ends_at: str


@dataclass(frozen=True)
class StudentRecord:
    student_code: str
    full_name: str
    birth_year: int | None
    phone: str | None


@dataclass(frozen=True)
class EnrollmentRecord:
    class_key: str
    student_code: str


TIME_RE = re.compile(r"(\d{1,2})[:.](\d{2})\s*[-–]\s*(\d{1,2})[:.](\d{2})")


def clean(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def slugify(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    ascii_text = normalized.encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^A-Za-z0-9]+", "-", ascii_text).strip("-").upper() or "UNKNOWN"


def sql_string(value: object | None) -> str:
    if value is None:
        return "null"
    return "'" + str(value).replace("'", "''") + "'"


def sql_int(value: int | None) -> str:
    return "null" if value is None else str(value)


def values_rows(rows: list[tuple[str, ...]]) -> str:
    return ",\n".join("  (" + ", ".join(row) + ")" for row in rows)


def normalize_time(value: str) -> tuple[str, str] | None:
    match = TIME_RE.search(value)
    if not match:
        return None
    starts_at = f"{int(match.group(1)):02d}:{match.group(2)}"
    ends_at = f"{int(match.group(3)):02d}:{match.group(4)}"
    return starts_at, ends_at


def parse_teacher_room(value: object) -> tuple[str, str] | None:
    text = clean(value)
    match = re.match(r"^(\d{2,4})\s*-?\s*(.+?)$", text)
    if not match:
        return None
    room = match.group(1)
    teacher = clean(match.group(2)).upper()
    if not re.search(r"[A-Za-zÇĞİÖŞÜçğıöşü]", teacher):
        return None
    return teacher, room


def parse_birth_year(value: object) -> int | None:
    if value is None:
        return None
    if isinstance(value, int):
        return value if 1900 <= value <= 2100 else None
    if isinstance(value, float):
        year = int(value)
        return year if 1900 <= year <= 2100 else None
    match = re.search(r"(19|20)\d{2}", str(value))
    return int(match.group(0)) if match else None


def normalize_phone(value: object) -> str | None:
    phone = re.sub(r"\D+", "", str(value or ""))
    return phone or None


def looks_like_student_name(value: object) -> bool:
    text = clean(value)
    if len(text) < 3:
        return False
    return bool(re.search(r"[A-Za-zÇĞİÖŞÜçğıöşü]", text))


def find_target_sheet(wb):
    matches = [
        ws
        for ws in wb.worksheets
        if ws.sheet_state == "visible" and ws.title.startswith(TARGET_VISIBLE_SHEET_PREFIX)
    ]
    if len(matches) != 1:
        visible = [ws.title for ws in wb.worksheets if ws.sheet_state == "visible"]
        raise RuntimeError(
            f"Expected exactly one visible sheet starting with {TARGET_VISIBLE_SHEET_PREFIX!r}. "
            f"Found {len(matches)}. Visible sheets: {visible}"
        )
    return matches[0]


def collect_records(ws):
    session_rows = []
    for row_idx in range(1, ws.max_row + 1):
        row_text = " ".join(clean(ws.cell(row_idx, col).value) for col in range(1, ws.max_column + 1))
        time_value = normalize_time(row_text)
        if time_value:
            session_rows.append((row_idx, *time_value))

    classes: dict[str, ClassRecord] = {}
    students: dict[str, StudentRecord] = {}
    enrollments: set[EnrollmentRecord] = set()

    for index, (header_row, starts_at, ends_at) in enumerate(session_rows):
        next_header_row = session_rows[index + 1][0] if index + 1 < len(session_rows) else ws.max_row + 1
        teacher_row = header_row + 3
        student_header_row = header_row + 4
        data_start_row = header_row + 5
        data_end_row = next_header_row - 1

        for col in range(1, ws.max_column + 1):
            parsed = parse_teacher_room(ws.cell(teacher_row, col).value)
            if not parsed:
                continue

            teacher_name, room = parsed
            name_col = col + 1
            birth_col = col + 2
            phone_col = col + 3
            header_text = clean(ws.cell(student_header_row, name_col).value).casefold()
            if "ad/soyad" not in header_text:
                continue

            teacher_code = f"YAZ-{slugify(teacher_name)}"
            class_key = f"VISIBLE-{starts_at}-{teacher_code}-ROOM-{room}"
            classes[class_key] = ClassRecord(
                class_key=class_key,
                teacher_code=teacher_code,
                teacher_name=teacher_name,
                room=room,
                starts_at=starts_at,
                ends_at=ends_at,
            )

            for row_idx in range(data_start_row, data_end_row + 1):
                full_name = clean(ws.cell(row_idx, name_col).value)
                if not looks_like_student_name(full_name):
                    continue

                birth_year = parse_birth_year(ws.cell(row_idx, birth_col).value)
                phone = normalize_phone(ws.cell(row_idx, phone_col).value)
                stable_suffix = phone[-4:] if phone else f"R{row_idx}"
                student_code = f"YAZ-VISIBLE-{slugify(full_name)}-{birth_year or 'NOYEAR'}-{stable_suffix}"

                students[student_code] = StudentRecord(
                    student_code=student_code,
                    full_name=full_name,
                    birth_year=birth_year,
                    phone=phone,
                )
                enrollments.add(EnrollmentRecord(class_key=class_key, student_code=student_code))

    return (
        sorted(classes.values(), key=lambda item: (item.starts_at, item.room, item.teacher_name)),
        sorted(students.values(), key=lambda item: item.student_code),
        sorted(enrollments, key=lambda item: (item.class_key, item.student_code)),
    )


def build_sql(classes: list[ClassRecord], students: list[StudentRecord], enrollments: list[EnrollmentRecord]) -> str:
    summer_school_start_date = "2026-07-06"
    summer_school_end_date = "2026-08-12"
    teacher_rows = [
        (
            sql_string(teacher_code),
            sql_string(teacher_name),
            sql_string("Imported from visible yazokulu.xlsx worksheet"),
        )
        for teacher_code, teacher_name in sorted({(c.teacher_code, c.teacher_name) for c in classes})
    ]
    class_rows = [
        (
            sql_string(item.class_key),
            sql_string(item.teacher_code),
            sql_string(item.teacher_name),
            sql_string(item.room),
            sql_string(item.starts_at),
            sql_string(item.ends_at),
        )
        for item in classes
    ]
    student_rows = [
        (
            sql_string(item.student_code),
            sql_string(item.full_name),
            sql_int(item.birth_year),
            sql_string(item.phone),
        )
        for item in students
    ]
    enrollment_rows = [
        (sql_string(item.class_key), sql_string(item.student_code)) for item in enrollments
    ]

    return f"""-- Final approved import from the visible yazokulu.xlsx worksheet only.
-- Source sheet: YAZ OKULU-SINIF PLANLANMASI(TAS
-- Hidden worksheets are intentionally ignored.
-- Imports exactly 4 teachers, 7 sessions/classes, and 84 students from the approved preview.
-- Safe to re-run.

begin;

-- Business start date for this approved summer school import.
-- Do not use current_date here; the import may be run after classes started.
create temp table tmp_visible_yazokulu_dates (
  start_date date not null,
  end_date date not null
) on commit drop;

insert into tmp_visible_yazokulu_dates (start_date, end_date)
values (date '{summer_school_start_date}', date '{summer_school_end_date}');

with teacher_data(employee_code, display_name, bio) as (
  values
{values_rows(teacher_rows)}
)
insert into public.teachers (
  employee_code,
  display_name,
  bio,
  languages,
  is_active,
  hired_at
)
select
  employee_code,
  display_name,
  bio,
  array['English'],
  true,
  current_date
from teacher_data
on conflict (employee_code) do update
set
  display_name = excluded.display_name,
  bio = excluded.bio,
  languages = excluded.languages,
  is_active = excluded.is_active;

with student_data(student_code, full_name, birth_year, phone) as (
  values
{values_rows(student_rows)}
)
insert into public.students (
  student_code,
  full_name,
  phone,
  date_of_birth,
  native_language,
  target_language,
  current_level,
  enrollment_date,
  is_active,
  notes
)
select
  student_code,
  full_name,
  phone,
  case when birth_year is null then null else make_date(birth_year, 1, 1) end,
  'Turkish',
  'English',
  'beginner',
  current_date,
  true,
  concat_ws('; ', 'Imported from visible yazokulu.xlsx worksheet', case when birth_year is null then null else 'Birth year: ' || birth_year::text end)
from student_data
on conflict (student_code) do update
set
  full_name = excluded.full_name,
  phone = excluded.phone,
  date_of_birth = excluded.date_of_birth,
  native_language = excluded.native_language,
  target_language = excluded.target_language,
  current_level = excluded.current_level,
  is_active = excluded.is_active,
  notes = excluded.notes;

do $import$
declare
  row_data record;
  teacher_uuid uuid;
  existing_class_id uuid;
  class_name text;
begin
  for row_data in
    select
      c.class_key,
      c.teacher_code,
      c.teacher_name,
      c.room,
      c.starts_at::time as starts_at,
      c.ends_at::time as ends_at,
      t.id as teacher_id
    from (
      values
{values_rows(class_rows)}
    ) as c(class_key, teacher_code, teacher_name, room, starts_at, ends_at)
    join public.teachers t on t.employee_code = c.teacher_code
  loop
    teacher_uuid := row_data.teacher_id;
    class_name := 'American Life Summer School - ' || to_char(row_data.starts_at, 'HH24:MI') || '-' || to_char(row_data.ends_at, 'HH24:MI') || ' - ' || row_data.teacher_name || ' - Room ' || row_data.room;

    select id
    into existing_class_id
    from public.classes
    where name = class_name
      and teacher_id = teacher_uuid
    order by created_at desc
    limit 1;

    if existing_class_id is null then
      insert into public.classes (
        teacher_id,
        name,
        language,
        level,
        status,
        capacity,
        start_date,
        end_date,
        schedule,
        location,
        meeting_url
      )
      values (
        teacher_uuid,
        class_name,
        'English',
        'beginner',
        'active',
        16,
        (select start_date from tmp_visible_yazokulu_dates),
        (select end_date from tmp_visible_yazokulu_dates),
        jsonb_build_array(jsonb_build_object(
          'days', 'Monday-Wednesday',
          'starts_at', to_char(row_data.starts_at, 'HH24:MI'),
          'ends_at', to_char(row_data.ends_at, 'HH24:MI'),
          'source_sheet', 'YAZ OKULU-SINIF PLANLANMASI(TAS'
        )),
        'Room ' || row_data.room,
        null
      )
      returning id into existing_class_id;
    else
      update public.classes
      set
        language = 'English',
        level = 'beginner',
        status = 'active',
        capacity = 16,
        start_date = (select start_date from tmp_visible_yazokulu_dates),
        end_date = (select end_date from tmp_visible_yazokulu_dates),
        schedule = jsonb_build_array(jsonb_build_object(
          'days', 'Monday-Wednesday',
          'starts_at', to_char(row_data.starts_at, 'HH24:MI'),
          'ends_at', to_char(row_data.ends_at, 'HH24:MI'),
          'source_sheet', 'YAZ OKULU-SINIF PLANLANMASI(TAS'
        )),
        location = 'Room ' || row_data.room,
        meeting_url = null
      where id = existing_class_id;
    end if;

    insert into public.lessons (
      class_id,
      teacher_id,
      lesson_date,
      starts_at,
      ends_at,
      title,
      objectives,
      materials,
      homework,
      status
    )
    select
      existing_class_id,
      teacher_uuid,
      lesson_day::date,
      row_data.starts_at,
      row_data.ends_at,
      'Summer School Session - ' || row_data.teacher_name || ' Room ' || row_data.room,
      'Imported from visible yazokulu.xlsx worksheet.',
      null,
      null,
      'scheduled'
    from generate_series(
      (select start_date from tmp_visible_yazokulu_dates),
      (select end_date from tmp_visible_yazokulu_dates),
      interval '1 day'
    ) as lesson_days(lesson_day)
    where extract(isodow from lesson_day) in (1, 2, 3)
    on conflict (class_id, lesson_date, starts_at) do update
    set
      teacher_id = excluded.teacher_id,
      ends_at = excluded.ends_at,
      title = excluded.title,
      objectives = excluded.objectives,
      status = excluded.status;
  end loop;
end $import$;

with
  enrollment_data(class_key, student_code) as (
    values
{values_rows(enrollment_rows)}
  ),
  class_data(class_key, teacher_code, teacher_name, room, starts_at, ends_at) as (
    values
{values_rows(class_rows)}
  ),
  resolved_classes as (
    select
      c.class_key,
      cls.id as class_id
    from class_data c
    join public.teachers t on t.employee_code = c.teacher_code
    join public.classes cls
      on cls.teacher_id = t.id
      and cls.name = 'American Life Summer School - ' || to_char(c.starts_at::time, 'HH24:MI') || '-' || to_char(c.ends_at::time, 'HH24:MI') || ' - ' || c.teacher_name || ' - Room ' || c.room
  )
insert into public.class_students (
  class_id,
  student_id,
  status,
  joined_at
)
select
  rc.class_id,
  s.id,
  'active',
  (select start_date from tmp_visible_yazokulu_dates)
from enrollment_data e
join resolved_classes rc on rc.class_key = e.class_key
join public.students s on s.student_code = e.student_code
on conflict on constraint class_students_pkey do update
set
  status = excluded.status,
  joined_at = excluded.joined_at,
  left_at = null;

commit;
"""


def main() -> None:
    wb = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
    ws = find_target_sheet(wb)
    classes, students, enrollments = collect_records(ws)

    if len(classes) != 7:
        raise RuntimeError(f"Expected 7 visible-sheet sessions/classes, found {len(classes)}.")
    if len({item.teacher_name for item in classes}) != 4:
        raise RuntimeError("Expected exactly 4 visible-sheet teachers.")
    if len(students) != 84:
        raise RuntimeError(f"Expected 84 visible-sheet students, found {len(students)}.")
    if len(enrollments) != 84:
        raise RuntimeError(f"Expected 84 visible-sheet enrollments, found {len(enrollments)}.")

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(build_sql(classes, students, enrollments), encoding="utf-8")

    summary_lines = [
        "Approved visible worksheet import",
        f"Source sheet: {ws.title}",
        "Hidden worksheets: ignored",
        f"Teachers: {len({item.teacher_name for item in classes})}",
        f"Classes/sessions: {len(classes)}",
        f"Students: {len(students)}",
        "",
        "Classes:",
    ]
    for item in classes:
        count = sum(1 for enrollment in enrollments if enrollment.class_key == item.class_key)
        summary_lines.append(
            f"- {item.starts_at}-{item.ends_at} | Room {item.room} | {item.teacher_name} | {count} students"
        )

    SUMMARY_PATH.write_text("\n".join(summary_lines) + "\n", encoding="utf-8")
    print("\n".join(summary_lines))
    print(f"\nGenerated: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
