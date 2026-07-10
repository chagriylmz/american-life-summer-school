from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import json
import re
import unicodedata

import openpyxl


PROJECT_ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = PROJECT_ROOT / "yazokulu.xlsx"
APPROVED_IMPORT_SQL = PROJECT_ROOT / "sql" / "import_visible_yazokulu_approved.sql"
OUTPUT_SQL = PROJECT_ROOT / "sql" / "update_parent_phones_from_yazokulu.sql"
OUTPUT_SUMMARY = PROJECT_ROOT / "sql" / "update_parent_phones_from_yazokulu_summary.json"
TARGET_VISIBLE_SHEET_PREFIX = "YAZ OKULU-SINIF PLANLANMASI(TAS"


@dataclass(frozen=True)
class PhoneUpdate:
    student_code: str
    full_name: str
    birth_year: int | None
    row_idx: int
    phone: str | None


TIME_RE = re.compile(r"(\d{1,2})[:.](\d{2})\s*[-–]\s*(\d{1,2})[:.](\d{2})")


def clean(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def slugify(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    ascii_text = normalized.encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^A-Za-z0-9]+", "-", ascii_text).strip("-").upper() or "UNKNOWN"


def normalize_time(value: str) -> tuple[str, str] | None:
    match = TIME_RE.search(value)
    if not match:
        return None
    starts_at = f"{int(match.group(1)):02d}:{match.group(2)}"
    ends_at = f"{int(match.group(3)):02d}:{match.group(4)}"
    return starts_at, ends_at


def parse_teacher_room(value: object) -> tuple[str, str] | None:
    text = clean(value)
    match = re.match(r"^(\d{2,4})\s*-?\s*(.+?)$", text)
    if not match:
        return None
    room = match.group(1)
    teacher = clean(match.group(2)).upper()
    if not re.search(r"[A-Za-zÇĞİÖŞÜçğıöşü]", teacher):
        return None
    return teacher, room


def parse_birth_year(value: object) -> int | None:
    if value is None:
        return None
    if isinstance(value, int):
        return value if 1900 <= value <= 2100 else None
    if isinstance(value, float):
        year = int(value)
        return year if 1900 <= year <= 2100 else None
    match = re.search(r"(19|20)\d{2}", str(value))
    return int(match.group(0)) if match else None


def normalize_phone(value: object) -> str | None:
    digits = re.sub(r"\D+", "", str(value or ""))
    if not digits:
        return None
    if len(digits) == 10:
        return f"0{digits}"
    return digits


def looks_like_student_name(value: object) -> bool:
    text = clean(value)
    if len(text) < 3:
        return False
    return bool(re.search(r"[A-Za-zÇĞİÖŞÜçğıöşü]", text))


def find_target_sheet(workbook: openpyxl.Workbook):
    matches = [
        ws
        for ws in workbook.worksheets
        if ws.sheet_state == "visible" and ws.title.startswith(TARGET_VISIBLE_SHEET_PREFIX)
    ]
    if len(matches) != 1:
        visible = [ws.title for ws in workbook.worksheets if ws.sheet_state == "visible"]
        raise RuntimeError(
            f"Expected exactly one visible sheet starting with {TARGET_VISIBLE_SHEET_PREFIX!r}. "
            f"Found {len(matches)}. Visible sheets: {visible}"
        )
    return matches[0]


def extract_existing_student_codes() -> dict[tuple[str, int | None, int], str]:
    sql = APPROVED_IMPORT_SQL.read_text(encoding="utf-8")
    pattern = re.compile(
        r"\('(?P<code>YAZ-VISIBLE-[^']+)',\s*'(?P<name>(?:''|[^'])*)',\s*(?P<birth>null|\d+),\s*(?:null|'(?:''|[^']*)')\)",
        re.IGNORECASE,
    )
    mapping: dict[tuple[str, int | None, int], str] = {}
    occurrence: dict[tuple[str, int | None], int] = {}

    for match in pattern.finditer(sql):
        name = match.group("name").replace("''", "'")
        birth = None if match.group("birth").lower() == "null" else int(match.group("birth"))
        base_key = (slugify(name), birth)
        index = occurrence.get(base_key, 0) + 1
        occurrence[base_key] = index
        mapping[(base_key[0], base_key[1], index)] = match.group("code")

    if not mapping:
        raise RuntimeError(f"Could not extract existing student codes from {APPROVED_IMPORT_SQL}")
    return mapping


def collect_phone_updates() -> tuple[list[PhoneUpdate], int, int]:
    workbook = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
    ws = find_target_sheet(workbook)
    code_mapping = extract_existing_student_codes()
    seen: dict[tuple[str, int | None], int] = {}
    updates: list[PhoneUpdate] = []
    blank_numbers = 0
    not_found = 0

    session_rows: list[tuple[int, str, str]] = []
    for row_idx in range(1, ws.max_row + 1):
        row_text = " ".join(clean(ws.cell(row_idx, col).value) for col in range(1, ws.max_column + 1))
        time_value = normalize_time(row_text)
        if time_value:
            session_rows.append((row_idx, *time_value))

    for index, (header_row, _starts_at, _ends_at) in enumerate(session_rows):
        next_header_row = session_rows[index + 1][0] if index + 1 < len(session_rows) else ws.max_row + 1
        teacher_row = header_row + 3
        student_header_row = header_row + 4
        data_start_row = header_row + 5
        data_end_row = next_header_row - 1

        for col in range(1, ws.max_column + 1):
            if not parse_teacher_room(ws.cell(teacher_row, col).value):
                continue

            name_col = col + 1
            birth_col = col + 2
            phone_col = col + 3
            header_text = clean(ws.cell(student_header_row, name_col).value).casefold()
            if "ad/soyad" not in header_text:
                continue

            for row_idx in range(data_start_row, data_end_row + 1):
                full_name = clean(ws.cell(row_idx, name_col).value)
                if not looks_like_student_name(full_name):
                    continue

                birth_year = parse_birth_year(ws.cell(row_idx, birth_col).value)
                base_key = (slugify(full_name), birth_year)
                occurrence_index = seen.get(base_key, 0) + 1
                seen[base_key] = occurrence_index
                student_code = code_mapping.get((base_key[0], base_key[1], occurrence_index))
                if not student_code:
                    not_found += 1
                    continue

                phone = normalize_phone(ws.cell(row_idx, phone_col).value)
                if phone is None:
                    blank_numbers += 1
                    continue

                updates.append(
                    PhoneUpdate(
                        student_code=student_code,
                        full_name=full_name,
                        birth_year=birth_year,
                        row_idx=row_idx,
                        phone=phone,
                    )
                )

    return updates, blank_numbers, not_found


def sql_string(value: object | None) -> str:
    if value is None:
        return "null"
    return "'" + str(value).replace("'", "''") + "'"


def build_sql(updates: list[PhoneUpdate], blank_numbers: int, not_found: int) -> str:
    values = ",\n".join(
        f"  ({sql_string(item.student_code)}, {sql_string(item.phone)})"
        for item in sorted(updates, key=lambda update: update.student_code)
    )
    return f"""-- Update parent phone numbers from the latest visible yazokulu.xlsx worksheet.
-- Safe scope:
--   - Updates existing students only by public.students.student_code.
--   - Does not create students.
--   - Does not modify names, classes, teachers, attendance, lessons, or notes.
--   - Blank Excel phone numbers are skipped.
--   - The schema has guardian_phone for parent phone; no parent_phone column exists in the checked schema.

begin;

with phone_updates(student_code, parent_phone) as (
values
{values}
),
updated as (
  update public.students s
  set guardian_phone = u.parent_phone
  from phone_updates u
  where s.student_code = u.student_code
    and s.guardian_phone is distinct from u.parent_phone
  returning s.student_code
),
missing as (
  select u.student_code
  from phone_updates u
  left join public.students s on s.student_code = u.student_code
  where s.id is null
)
select
  (select count(*) from phone_updates) as total_students_processed,
  (select count(*) from updated) as phones_updated,
  {blank_numbers}::int as blank_numbers_skipped,
  (select count(*) from missing) + {not_found}::int as students_not_found;

commit;
"""


def main() -> None:
    updates, blank_numbers, not_found = collect_phone_updates()
    OUTPUT_SQL.write_text(build_sql(updates, blank_numbers, not_found), encoding="utf-8")
    OUTPUT_SUMMARY.write_text(
        json.dumps(
            {
                "source_workbook": str(WORKBOOK_PATH),
                "target_sql": str(OUTPUT_SQL),
                "total_students_processed": len(updates),
                "phones_update_candidates": len(updates),
                "blank_numbers_skipped": blank_numbers,
                "students_not_found_before_sql": not_found,
                "target_column": "public.students.guardian_phone",
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    print(OUTPUT_SUMMARY.read_text(encoding="utf-8"))


if __name__ == "__main__":
    main()
