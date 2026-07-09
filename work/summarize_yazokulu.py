from pathlib import Path
import re
import openpyxl

wb = openpyxl.load_workbook(Path("yazokulu.xlsx"), data_only=True)
ws = wb[next(name for name in wb.sheetnames if "YAZ OKULU 1 TEMMUZ" in name)]

time_re = re.compile(r"(\d{1,2})[:.](\d{2})\s*[-–]\s*(\d{1,2})[:.](\d{2})")
for row in range(1, ws.max_row + 1):
    vals = []
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row, col).value
        if v is not None and str(v).strip():
            vals.append((col, str(v).strip().replace("\n", " ")))
    if any(time_re.search(v) for _, v in vals):
        print("TIME ROW", row, vals)
        for check_row in range(row + 1, min(row + 7, ws.max_row + 1)):
            row_vals = []
            for col in range(1, 18):
                v = ws.cell(check_row, col).value
                if v is not None and str(v).strip():
                    row_vals.append((col, str(v).strip().replace("\n", " ")))
            print(" ", check_row, row_vals)
