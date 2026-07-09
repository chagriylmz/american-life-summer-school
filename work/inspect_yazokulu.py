from pathlib import Path
import json
import openpyxl

path = Path("yazokulu.xlsx")
wb = openpyxl.load_workbook(path, data_only=True)
print("sheets:", wb.sheetnames)
target = next((name for name in wb.sheetnames if "YAZ OKULU 1 TEMMUZ" in name), wb.sheetnames[0])
ws = wb[target]
print("active_sheet:", ws.title)
print("dimensions:", ws.max_row, ws.max_column)
print("merged_ranges:", [str(rng) for rng in list(ws.merged_cells.ranges)[:80]])

non_empty = []
for row in ws.iter_rows():
    for cell in row:
        value = cell.value
        if value is not None and str(value).strip() != "":
            non_empty.append(
                {
                    "cell": cell.coordinate,
                    "row": cell.row,
                    "col": cell.column,
                    "value": str(value).strip(),
                }
            )

print("non_empty_count:", len(non_empty))
print("first_250_non_empty:")
for item in non_empty[:250]:
    print(json.dumps(item, ensure_ascii=False))

print("time_matches:")
for item in non_empty:
    value = item["value"]
    if "13" in value or "15" in value or "17" in value:
        print(json.dumps(item, ensure_ascii=False))
