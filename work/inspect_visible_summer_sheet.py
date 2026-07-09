from pathlib import Path
import json
import openpyxl

wb = openpyxl.load_workbook(Path("yazokulu.xlsx"), data_only=True)
ws = next(
    sheet
    for sheet in wb.worksheets
    if sheet.sheet_state == "visible" and sheet.title.startswith("YAZ OKULU-SINIF PLANLANMASI(TAS")
)

print("sheet:", ws.title, ws.sheet_state, ws.max_row, ws.max_column)
for row_idx in range(1, min(ws.max_row, 120) + 1):
    cells = []
    for col in range(1, min(ws.max_column, 30) + 1):
        value = ws.cell(row_idx, col).value
        if value is not None and str(value).strip():
            cells.append({"cell": ws.cell(row_idx, col).coordinate, "value": str(value).strip()})
    if cells:
        print(json.dumps({"row": row_idx, "cells": cells}, ensure_ascii=False))
